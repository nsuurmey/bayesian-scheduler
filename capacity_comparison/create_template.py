#!/usr/bin/env python3
"""
Generate a pre-filled Excel template for capacity planning input.

Creates portfolio_template.xlsx with three sheets:
  - Projects: one row per project
  - Tasks: one row per task (linked to projects)
  - Config: team capacity and historical data

Run:
    python create_template.py
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ── Styling ──────────────────────────────────────────────────────────────
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="CBD5E0"),
    right=Side(style="thin", color="CBD5E0"),
)
WRAP = Alignment(wrap_text=True, vertical="top")


def _style_header(ws, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 30)


def create_template(path="portfolio_template.xlsx"):
    wb = openpyxl.Workbook()

    # ── Sheet 1: Projects ─────────────────────────────────────────────
    ws_proj = wb.active
    ws_proj.title = "Projects"

    proj_headers = [
        "project_name",
        "type",
        "value",
        "mandatory",
        "dependencies",
        "optimistic_weeks",
        "likely_weeks",
        "pessimistic_weeks",
        "earliest_start_month",
        "probability",
        "ds_month_1", "ds_month_2", "ds_month_3", "ds_month_4", "ds_month_5", "ds_month_6",
        "geo_month_1", "geo_month_2", "geo_month_3", "geo_month_4", "geo_month_5", "geo_month_6",
        "pm_month_1", "pm_month_2", "pm_month_3", "pm_month_4", "pm_month_5", "pm_month_6",
    ]
    for c, h in enumerate(proj_headers, 1):
        ws_proj.cell(row=1, column=c, value=h)

    # Sample data (matches shared_dataset.py)
    sample_projects = [
        ["Seismic Reprocessing Platform", "large", 95, "yes", "",
         16, 22, 32, 0, 1.0,
         3, 4, 4, 3, 2, 2,
         2, 2, 3, 2, 1, 0,
         1, 1, 1, 1, 1, 1],
        ["Well-Log ML Predictor", "medium", 80, "yes", "",
         10, 14, 20, 0, 1.0,
         2, 3, 3, 2, "", "",
         1, 1, 1, 0, "", "",
         1, 1, 1, 1, "", ""],
        ["Reservoir Simulation Dashboard", "large", 90, "yes", "Seismic Reprocessing Platform",
         14, 20, 28, 5, 1.0,
         2, 3, 4, 3, 2, "",
         2, 2, 2, 1, 1, "",
         1, 1, 1, 1, 1, ""],
        ["Automated Core Analysis", "medium", 75, "no", "",
         8, 12, 18, 2, 0.85,
         2, 2, 2, "", "", "",
         2, 3, 2, "", "", "",
         1, 1, 0, "", "", ""],
        ["Production Forecasting Upgrade", "small", 65, "yes", "",
         5, 8, 12, 1, 1.0,
         2, 2, "", "", "", "",
         1, 1, "", "", "", "",
         1, 1, "", "", "", ""],
        ["Geospatial Data Lake", "large", 85, "no", "",
         18, 26, 36, 0, 0.70,
         3, 4, 4, 3, 3, 2,
         1, 1, 2, 2, 1, 1,
         1, 1, 1, 1, 1, 1],
        ["Environmental Compliance Tracker", "small", 70, "yes", "Well-Log ML Predictor",
         4, 6, 10, 4, 1.0,
         1, 2, "", "", "", "",
         1, 1, "", "", "", "",
         1, 1, "", "", "", ""],
    ]

    for r, row_data in enumerate(sample_projects, 2):
        for c, val in enumerate(row_data, 1):
            ws_proj.cell(row=r, column=c, value=val)

    _style_header(ws_proj, len(proj_headers))
    _auto_width(ws_proj)
    ws_proj.freeze_panes = "A2"

    # ── Sheet 2: Tasks ────────────────────────────────────────────────
    ws_tasks = wb.create_sheet("Tasks")

    task_headers = [
        "project_name",
        "task_id",
        "task_name",
        "optimistic_weeks",
        "likely_weeks",
        "pessimistic_weeks",
        "predecessors",
        "data_scientists",
        "geoscientists",
        "project_managers",
    ]
    for c, h in enumerate(task_headers, 1):
        ws_tasks.cell(row=1, column=c, value=h)

    sample_tasks = [
        # Seismic Reprocessing Platform
        ["Seismic Reprocessing Platform", "S1", "Requirements & Scoping", 2, 3, 5, "", 1, 2, 0],
        ["Seismic Reprocessing Platform", "S2", "Data Ingestion Pipeline", 3, 5, 8, "S1", 3, 0, 0],
        ["Seismic Reprocessing Platform", "S3", "Processing Algorithm Dev", 4, 6, 10, "S1", 2, 2, 0],
        ["Seismic Reprocessing Platform", "S4", "Integration & Testing", 3, 4, 6, "S2, S3", 2, 1, 0],
        ["Seismic Reprocessing Platform", "S5", "Deployment & Handover", 2, 3, 5, "S4", 1, 0, 1],
        # Well-Log ML Predictor
        ["Well-Log ML Predictor", "W1", "Data Audit & Cleaning", 2, 3, 5, "", 1, 1, 0],
        ["Well-Log ML Predictor", "W2", "Feature Engineering", 2, 3, 5, "W1", 2, 0, 0],
        ["Well-Log ML Predictor", "W3", "Model Training & Tuning", 3, 4, 6, "W2", 2, 0, 0],
        ["Well-Log ML Predictor", "W4", "Validation & Reporting", 2, 3, 4, "W3", 1, 1, 1],
        # Reservoir Simulation Dashboard
        ["Reservoir Simulation Dashboard", "R1", "Dashboard Architecture", 2, 3, 4, "", 2, 0, 1],
        ["Reservoir Simulation Dashboard", "R2", "Simulation Engine Wrapper", 4, 6, 9, "R1", 3, 2, 0],
        ["Reservoir Simulation Dashboard", "R3", "Visualization Layer", 3, 4, 6, "R1", 2, 0, 0],
        ["Reservoir Simulation Dashboard", "R4", "User Testing & Polish", 3, 5, 8, "R2, R3", 1, 1, 1],
        # Automated Core Analysis
        ["Automated Core Analysis", "C1", "Image Pipeline Setup", 2, 3, 5, "", 1, 1, 0],
        ["Automated Core Analysis", "C2", "Classification Model", 3, 4, 6, "C1", 2, 1, 0],
        ["Automated Core Analysis", "C3", "Integration & QA", 2, 3, 5, "C2", 1, 2, 0],
        # Production Forecasting Upgrade
        ["Production Forecasting Upgrade", "P1", "Current Model Audit", 1, 2, 3, "", 1, 1, 0],
        ["Production Forecasting Upgrade", "P2", "Algorithm Improvements", 2, 3, 5, "P1", 2, 0, 0],
        ["Production Forecasting Upgrade", "P3", "Back-testing & Deploy", 1, 2, 3, "P2", 1, 0, 1],
        # Geospatial Data Lake
        ["Geospatial Data Lake", "G1", "Cloud Architecture Design", 3, 4, 7, "", 2, 0, 1],
        ["Geospatial Data Lake", "G2", "Data Migration", 4, 7, 10, "G1", 3, 1, 0],
        ["Geospatial Data Lake", "G3", "API & Access Layer", 3, 5, 8, "G1", 2, 0, 0],
        ["Geospatial Data Lake", "G4", "Spatial Indexing Engine", 3, 5, 7, "G2", 2, 2, 0],
        ["Geospatial Data Lake", "G5", "Testing & Documentation", 2, 3, 5, "G3, G4", 1, 1, 1],
        # Environmental Compliance Tracker
        ["Environmental Compliance Tracker", "E1", "Regulatory Data Mapping", 1, 2, 3, "", 1, 1, 0],
        ["Environmental Compliance Tracker", "E2", "Alert & Reporting Module", 2, 3, 5, "E1", 2, 0, 1],
        ["Environmental Compliance Tracker", "E3", "Stakeholder Review", 1, 1, 2, "E2", 0, 0, 1],
    ]

    for r, row_data in enumerate(sample_tasks, 2):
        for c, val in enumerate(row_data, 1):
            ws_tasks.cell(row=r, column=c, value=val)

    _style_header(ws_tasks, len(task_headers))
    _auto_width(ws_tasks)
    ws_tasks.freeze_panes = "A2"

    # ── Sheet 3: Config ───────────────────────────────────────────────
    ws_config = wb.create_sheet("Config")

    # Team capacity section
    config_data = [
        ["TEAM CAPACITY", "", ""],
        ["Role", "Headcount", ""],
        ["data_scientists", 8, ""],
        ["geoscientists", 5, ""],
        ["project_managers", 4, ""],
        ["", "", ""],
        ["PLANNING", "", ""],
        ["horizon_months", 12, ""],
        ["", "", ""],
        ["HISTORICAL DURATIONS (weeks)", "", ""],
        ["Category", "Durations (comma-separated)", ""],
        ["small", "6, 8, 7, 9, 7, 8, 10, 6", ""],
        ["medium", "12, 15, 14, 16, 13, 15", ""],
        ["large", "20, 24, 22, 26, 23, 28, 21", ""],
    ]

    for r, row_data in enumerate(config_data, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws_config.cell(row=r, column=c, value=val)

    # Style section headers
    for row_idx in [1, 7, 10]:
        cell = ws_config.cell(row=row_idx, column=1)
        cell.font = Font(bold=True, size=12, color="1A365D")

    for col in [1, 2]:
        ws_config.column_dimensions[get_column_letter(col)].width = 35

    # ── Save ──────────────────────────────────────────────────────────
    output = os.path.join(os.path.dirname(__file__), path)
    wb.save(output)
    print(f"Template saved to: {output}")
    return output


if __name__ == "__main__":
    create_template()
